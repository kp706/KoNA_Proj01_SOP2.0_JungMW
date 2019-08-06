
import xlrd
from openpyxl import load_workbook
import sys
import os
#from datetime import datetime
import io
#import numpy as np
from PyQt5.QtWidgets import (QApplication,QFileDialog, QTextEdit, QComboBox, QVBoxLayout,QWidget,  QPlainTextEdit,QPushButton, QDesktopWidget,QGridLayout, QLabel, QLineEdit,QRadioButton)
from PyQt5.QtGui import QIcon,QColor
from PyQt5.QtCore import QCoreApplication
import Validation

#excelPath = 'D:/minwoo/Working_Directory/03_이대엽_크로마틴 구조기반 간암 유방암 예후예측 3D-nucleome 바이오마커 발굴_20190710.xlsx'



excelPath = ""
'''
input value = excel File
'''
class KonaValidation(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()
    def center(self): #창을 가운데 띄우기위한 함수
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def searchExcelFileButtonClicked(self,path,errbox):
        fname = QFileDialog.getOpenFileName(self)
        path.setText(fname[0])
        errbox.clear()


    def initUI(self):

        self.resize(1000,600)
        self.center()
        self.setWindowTitle("Kona Validation Program")
        excelPathInput = QLineEdit(self)
        excelFileSearch = QPushButton("엑셀 파일 찾기",self)
        validation = QPushButton('Validation Start',self)
        errorBox = QTextEdit(self)

        grid = QGridLayout()
        self.setLayout(grid)
        grid.addWidget(QLabel('엑셀 파일:'),1,0)
        grid.addWidget(excelPathInput,1,1)
        grid.addWidget(excelFileSearch,1,2)
        grid.addWidget(validation,2,1)
        grid.addWidget(errorBox,3,1)
        excelFileSearch.clicked.connect(lambda: self.searchExcelFileButtonClicked(excelPathInput,errorBox))
        validation.clicked.connect(lambda: self.run(str(excelPathInput.text()),errorBox))
        self.show()




    bioSample_SampleName = []
    experiment_SampleName = []

    def run(self,excelPathInputValue,errbox):
        try:

            targetExcel = load_workbook(excelPathInputValue,data_only=True) # 엑셀 연다.

            bioProjectSheetName = ''
            bioSampleSheetName = []
            sampleTypeSheetName = []
            experimentSheetName = []

            sheets = targetExcel.sheetnames
            #시트 이름들을 가져와서 포함되는 단어에 따라서 각각의 배열에 추가
            for sheet in sheets:
                if 'BioProject' in str(sheet):
                    bioProjectSheetName += str(sheet)
                elif 'BioSample' in str(sheet):
                    bioSampleSheetName.append(str(sheet))
                elif 'Sample type' in str(sheet):
                    sampleTypeSheetName.append(str(sheet))
                elif 'Experiment' in str(sheet):
                    experimentSheetName.append(str(sheet))


            #biosample,sampletype,experiment 쌍의 개수만큼 반복한다.
            rotation = len(bioSampleSheetName)

            bioProject = targetExcel[bioProjectSheetName]
            Validation.bioProject_Validation(bioProject,bioProjectSheetName,errbox) #BioProject는 1개뿐이므로 그냥 validation 실행

            #나머지 시트들은 존재하는 개수만큼 실행한다.
            i = 0
            while i < rotation:
                bioSample = targetExcel[bioSampleSheetName[i]]
                sampleType = targetExcel[sampleTypeSheetName[i]]
                experiment = targetExcel[experimentSheetName[i]]

                Validation.bioSample_Validation(bioSample,sampleType,bioSampleSheetName[i],errbox)
                Validation.sampleType_Validation(sampleType,sampleTypeSheetName[i],errbox)
                Validation.Experiment_Validation(experiment,experimentSheetName[i],errbox)
                i += 1

        except IOError as err:
            errbox.insertPlainText("IO Error : " + str(err))




if __name__ == '__main__':

    app = QApplication(sys.argv) #어플리케이션 객체 생성
    ex = KonaValidation()
    #ex.show()
    sys.exit(app.exec_())

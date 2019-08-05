
import xlrd
from openpyxl import load_workbook
import sys
import os
from datetime import datetime
import io
import numpy as np
from PyQt5.QtWidgets import (QApplication,QFileDialog, QTextEdit, QComboBox, QVBoxLayout,QWidget,  QPlainTextEdit,QPushButton, QDesktopWidget,QGridLayout, QLabel, QLineEdit,QRadioButton)
from PyQt5.QtGui import QIcon,QColor
from PyQt5.QtCore import QCoreApplication
sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')


#excelPath = 'D:/minwoo/Working_Directory/03_이대엽_크로마틴 구조기반 간암 유방암 예후예측 3D-nucleome 바이오마커 발굴_20190710.xlsx'
redColor = QColor(255,0,0)
blueColor = QColor(0,0,255)
blackColor = QColor(0,0,0)


excelPath = ""
'''
input value = excel File
'''
class KonaValidation(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()
    def center(self): #창을 가운데 띄우기위한 함수
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def searchExcelFileButtonClicked(self,path,errbox):
        fname = QFileDialog.getOpenFileName(self)
        path.setText(fname[0])
        errbox.clear()


    def initUI(self):

        self.resize(1000,600)
        self.center()
        self.setWindowTitle("Kona Validation Program")
        excelPathInput = QLineEdit(self)
        excelFileSearch = QPushButton("엑셀 파일 찾기",self)
        validation = QPushButton('Validation Start',self)
        errorBox = QTextEdit(self)

        grid = QGridLayout()
        self.setLayout(grid)
        grid.addWidget(QLabel('엑셀 파일:'),1,0)
        grid.addWidget(excelPathInput,1,1)
        grid.addWidget(excelFileSearch,1,2)
        grid.addWidget(validation,2,1)
        grid.addWidget(errorBox,3,1)
        excelFileSearch.clicked.connect(lambda: self.searchExcelFileButtonClicked(excelPathInput,errorBox))
        validation.clicked.connect(lambda: self.run(str(excelPathInput.text()),errorBox))
        self.show()




    bioSample_SampleName = []
    experiment_SampleName = []

    def run(self,excelPathInputValue,errbox):
        try:

            targetExcel = load_workbook(excelPathInputValue,data_only=True) # 엑셀 연다.

            bioProjectSheetName = ''
            bioSampleSheetName = []
            sampleTypeSheetName = []
            experimentSheetName = []

            sheets = targetExcel.sheetnames
            #시트 이름들을 가져와서 포함되는 단어에 따라서 각각의 배열에 추가
            for sheet in sheets:
                if 'BioProject' in str(sheet):
                    bioProjectSheetName += str(sheet)
                elif 'BioSample' in str(sheet):
                    bioSampleSheetName.append(str(sheet))
                elif 'Sample type' in str(sheet):
                    sampleTypeSheetName.append(str(sheet))
                elif 'Experiment' in str(sheet):
                    experimentSheetName.append(str(sheet))


            #biosample,sampletype,experiment 쌍의 개수만큼 반복한다.
            rotation = len(bioSampleSheetName)

            bioProject = targetExcel[bioProjectSheetName]
            self.bioProject_Validation(bioProject,bioProjectSheetName,errbox) #BioProject는 1개뿐이므로 그냥 validation 실행

            #나머지 시트들은 존재하는 개수만큼 실행한다.
            i = 0
            while i < rotation:
                bioSample = targetExcel[bioSampleSheetName[i]]
                sampleType = targetExcel[sampleTypeSheetName[i]]
                experiment = targetExcel[experimentSheetName[i]]

                self.bioSample_Validation(bioSample,sampleType,bioSampleSheetName[i],errbox)
                self.sampleType_Validation(sampleType,sampleTypeSheetName[i],errbox)
                self.Experiment_Validation(experiment,experimentSheetName[i],errbox)
                i += 1

        except IOError as err:
            errbox.insertPlainText("IO Error : " + str(err))


    def checkingReleaseDate(self,release_date):
        '''
        날짜를 받아와서 현재 날짜를 기준으로 1년 이내이면 true, 이후이면 false 반환
        '''
        (year,month,day) = release_date.split('-',2)
        currentTime = datetime.now()
        day = str(day).split(' ',1)[0]
        releaseTime = datetime(int(str(year)),int(str(month)),int(str(day)))

        if int(((releaseTime-currentTime)).days) > 365:
            return False
        else:
            return True

    def notMatchedFieldName(self,sheet,index,fieldName):
        '''
        필드이름이 올바른지 확인한다. (true,false)
        '''
        if str(sheet[str(object=index)].value) != fieldName:
            return True
        else:
            return False;

    def bioProject_Validation(self,targetSheet,sheetName,errbox):
        '''
        bioProject 시트에서 조건들을 검사한다.
        '''
        flag = 0

        errbox.setTextColor(blackColor)
        if self.notMatchedFieldName(targetSheet,'A17','Submission date'):
            '''
            A17이 Submission date이 아니라면 에러 메시지를 출력한다.
            '''

            errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Submission date 필드의 위치가 템플릿 양식과 일치하지 않습니다." (17 row)\n')
            flag += 1
        else:
            if not str(object=targetSheet[str('E')+str(17)].value).count('-') == 2:
                errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Submission date" 입력값 형식이 적절하지 않습니다.(17 row, 입력형식:YYYY-MM-DD)\n')
                flag += 1

        #Release date Check
        if self.notMatchedFieldName(targetSheet,'A18','Release date selection'):
            '''
            A18이 Release date section이 아니라면 에러메시지를 출력한다.
            '''
            errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Release date selection 필드의 위치가 템플릿 양식과 일치하지 않습니다." (18 row)\n')
            flag += 1
        else:
            if str(targetSheet[str('E')+str(18)].value) == "Release on specified date":
                if not str(targetSheet[str('E')+str(19)].value).count('-') == 2:
                    errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Release on specified date" 를 선택한 경우 반드시 공개날짜를 입력해야합니다.(19 row,입력형식:YYYY-MM-DD)\n')
                    flag = 1
                elif not self.checkingReleaseDate(str(targetSheet[str('E')+str(19)].value)):
                    errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Release Date"가 현재로부터 1년 이후로 설정되어있습니다.(19 row)\n')
                    flag = 1
            elif not str(targetSheet[str('E')+str(18)].value) == "Release immediately following curation (recommended)":
                errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Release date section" 선택 입력값이 적절하지 않습니다.(18 row, 설명에있는 예시중 선택해야함)\n')
                flag += 1

        #M/O Field Check - M인 필드는 꼭 입력이 되어야한다.
        i = 3
        while i < 60:
            '''
            필수 입력값(M)이 입력되었는지 검사한다.
            '''
            if str(targetSheet['B'+str(i)].value)=='M':
                if str(targetSheet['E'+str(i)].value) == 'None' or str(targetSheet['E'+str(i)].value) == 'NA' or str(targetSheet['E'+str(i)].value) == 'NA':
                    errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Mandatory값(필수 입력값)이 입력되지 않았습니다." ('+ str(object=i) + ' row )\n')
                    flag += 1
            i += 1

        #Government department Check
        if self.notMatchedFieldName(targetSheet,'A21','Government department (국문)'):
            errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Government department 필드의 위치가 템플릿 양식과 일치하지 않습니다." (21 row)\n')
            flag += 1
        else:
            if str(object=targetSheet[str('E')+str(21)].value) not in ['과기정통부','해양수산부','보건복지부','농림축산부','산업부','농진청','산림청',]:
                errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Government department" 선택 입력 값이 잘못되었습니다. (21 row)\n')
                flag += 1


        #Project type Check
        if self.notMatchedFieldName(targetSheet,'A26','Project type'):
            errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Project type 필드의 위치가 템플릿 양식과 일치하지 않습니다." (26 row)\n')
            flag += 1
        else:
            if str(object=targetSheet[str('E')+str(26)].value)=='총괄':
                errbox.insertPlainText('[WARNING] ['+str(sheetName)+'] "Project type"이 총괄로 등록된 경우 따로 정리해서 결정해야합니다.(26 row)\n')
                flag += 1

        if(flag==0):
            errbox.setTextColor(blueColor)
            errbox.insertPlainText("<<< " + str(sheetName) + " : NO ERROR >>>\n")
        else:
            errbox.setTextColor(redColor)
            errbox.insertPlainText("<<< " + str(sheetName) + " : " + str(flag) + " ERROR >>>\n")





    def bioSample_Validation(self,targetSheet,compareSheet,sheetName,errbox):
        flag = 0
        errbox.setTextColor(blackColor)
        #Submission date check
        if self.notMatchedFieldName(targetSheet,'A19','Submission date'):
            errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Submission date 필드의 위치가 템플릿 양식과 일치하지 않습니다." (19 row)\n')
            flag += 1
        else:
            if not str(object=targetSheet[str('E')+str(19)].value).count('-') == 2:
                errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] Submission date 입력값 형식이 적절하지 않습니다.(19 row, 입력형식:YYYY-MM-DD)\n')
                flag += 1

        #Release date Check
        if str(targetSheet[str('E')+str(20)].value) == "Release on specified date":
            if not str(targetSheet[str('E')+str(21)].value).count('-') == 2:
                errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Release on specified date" 를 선택한 경우 반드시 공개날짜를 입력해야합니다.(21 row,입력형식:YYYY-MM-DD\n')
                flag += 1
            elif not self.checkingReleaseDate(str(targetSheet[str('E')+str(21)].value)):
                errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Release Date"가 현재로부터 1년 이후로 설정되어있습니다.(21 row)\n')
                flag += 1
        elif not str(targetSheet[str('E')+str(20)].value) == "Release immediately following curation (recommended)":
            errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Release date section" 선택 입력값이 적절하지 않습니다.(20 row, 설명에있는 예시중 선택해야함)\n')
            flag += 1

        #Project accession check
        if self.notMatchedFieldName(targetSheet,'A17','Project accession '):
            errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Project accession 필드의 위치가 템플릿 양식과 일치하지 않습니다." (17 row)\n')
            flag += 1
        else:
            if targetSheet[str('E')+str(17)].value == None or str(object=targetSheet[str('E')+str(17)].value) == 'NA' or  str(object=targetSheet[str('E')+str(17)].value) == 'N/A':
                errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Project accession" 값을 입력해야합니다.(17 row)\n')
                flag += 1

        if self.notMatchedFieldName(targetSheet,'A27','Sample type'):
            errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Sampletype type" 필드의 위치가 템플릿 양식과 일치하지 않습니다. (27 row)\n')
            flag += 1
        else:
            '''
            bio sample sheet의 sample type에 입력된 문자열의 키워드가 다음시트인 sample type 시트의 A1에 입력된 값과 일치하는값이 있는지 확인한다.
            '''
            temp1 = str(targetSheet['E27'].value).split(' ')
            temp2 = str(compareSheet['A1'].value).split(' ')
            included = False
            for item in temp1:
                if item in temp2:
                    included = True
                    break

            if not included:
                errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "bioSample의 sampletype"이 "SampleType" 시트에 있는 "sampletype" 값과 다릅니다." (27 row)\n')
                flag += 1

        if flag==0:
            errbox.setTextColor(blueColor)
            errbox.insertPlainText("<<< " + str(sheetName)+ " : NO ERROR >>>\n")
        else:
            errbox.setTextColor(redColor)
            errbox.insertPlainText("<<< " + str(sheetName)+ " : " + str(flag) + " ERROR >>>\n")



    def sampleType_Validation(self,targetSheet,sheetName,errbox):

        flag = 0
        i = 5
        errbox.setTextColor(blackColor)
        '''
        bioSample_SampleName 이라는 배열에 Sample Name을 다 넣고 중복이 있는지 본다.
        '''
        while True:
            temp = targetSheet['A'+str(i)].value
            if temp == None:
                break
            else:
                self.bioSample_SampleName.append(str(temp))
                i += 1
                #여기까지하면 배열에 이름들 저장
        nameSet = list(set(self.bioSample_SampleName))

        nameSet.sort()
        self.bioSample_SampleName.sort()

        if len(nameSet)!=len(self.bioSample_SampleName):
            errbox.insertPlainText('[ERROR] ['+sheetName+'] "sample type에 이름이 중복되는 데이터가 있습니다."\n')
            flag += 1

        i = 5
        column = ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W']
        duplicationCheckArr = []
        while True:
            if targetSheet['A'+str(i)].value == None:
                break
            temp = ""
            for col in column:
                temp += str(targetSheet[str(col)+str(i)].value).strip()
            duplicationCheckArr.append(temp)
            i += 1

        #무한루프를 빠져나오면 배열에 저장 완료
        duplicationCheckSet = list(set(duplicationCheckArr))
        if len(duplicationCheckArr)!=len(duplicationCheckSet):
            errbox.insertPlainText('[ERROR] ['+sheetName+'] "데이터 리스트 중에 이름을 제외한 모든 값이 중복되는 데이터 쌍이 있습니다."\n')
            flag += 1


        if flag==0:
            errbox.setTextColor(blueColor)
            errbox.insertPlainText("<<< "+str(sheetName)+ " : NO ERROR >>>\n")
        else:
            errbox.setTextColor(redColor)
            errbox.insertPlainText("<<< "+str(sheetName) +" : " + str(flag) + " ERROR >>>\n")




    def Experiment_Validation(self,targetSheet,sheetName,errbox):
        flag= 0
        errbox.setTextColor(blackColor)
        i = 5
        #Save Sample names
        while True:
            temp = object=targetSheet[str('A')+str(i)].value
            if temp == None:
                break
            else:
                self.experiment_SampleName.append(str(temp))
                i += 1
                #Experiment에 있는 sample name 배열에 저장

        nameSet = list(set(self.experiment_SampleName))
        compareNameSet = list(set(self.bioSample_SampleName))

        if len(self.experiment_SampleName) != len(nameSet):
            errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "sample name"에 이름이 중복되는 데이터가 있습니다.\n')
            flag += 1

        if len(set(nameSet)-set(compareNameSet))!=0:
            errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "sample name"에 있는 항목중에 "BioSample"에 없는 값이 있습니다."\n')
            flag +=1




        #Release date Check
        if str(targetSheet[str('C'+str(5))].value) == "Release on specified date":
            i = 5
            while True:
                temp = targetSheet[str('D')+str(i)].value
                if temp == None:
                    break
                else:
                    if not (self.checkingReleaseDate(temp)):
                        errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Release Date"가 현재로부터 1년 이후로 설정되어있습니다.(' + str(i) + ' row)\n')
                        flag += 1
                    i += 1



        #Size value check
        i = 5
        while True:
            temp = targetSheet[str('Q')+str(i)].value
            if temp == None:
                break
            if str(temp) == "Paired-end":
                if (str(targetSheet[str('R')+str(i)].value)=='None' or str(targetSheet[str('R')+str(i)].value)=='NA') and (str(targetSheet[str('S')+str(i)].value)=='None' or str(targetSheet[str('S')+str(i)].value)=='NA'):
                    errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] "Paired-end"인 경우 "Insert size" 또는 "Normal size"중 하나는 값을 입력해야합니다.('+str(i)+' row)\n')
                    flag += 1
                i += 1
            else:
                i += 1



        i = 5
        column = ['E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U']
        duplicationCheckArr = []
        while True:
            if targetSheet['A'+str(i)].value == None:
                break
            temp = ""
            for col in column:
                temp += str(targetSheet[str(col)+str(i)].value).strip()
            duplicationCheckArr.append(temp)
            i += 1
            #빠져나오면 uplicationCheckArr에 원소들 저장
        duplicationCheckSet = list(set(duplicationCheckArr))

        if not len(duplicationCheckArr)==len(duplicationCheckSet):
            errbox.insertPlainText('[ERROR] ['+str(sheetName)+'] E~U 열이 모두 중복되는 데이터가 있습니다.\n')
            flag += 1
            
        if flag == 0:
            errbox.setTextColor(blueColor)
            errbox.insertPlainText('<<< ' +str(sheetName) + ' : NO ERROR >>>\n')
        else:
            errbox.setTextColor(redColor)
            errbox.insertPlainText("<<< " + str(sheetName) + " : " + str(flag) + " ERROR >>>\n")



if __name__ == '__main__':

    app = QApplication(sys.argv) #어플리케이션 객체 생성
    ex = KonaValidation()
    #ex.show()
    sys.exit(app.exec_())
